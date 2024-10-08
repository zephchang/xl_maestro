import traceback
import re
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string

# CELLS
def extract_cells(formula):
    """
    Extract cell references from a formula.

    Args:
        formula (str): The formula to extract cell references from.

    Returns:
        cel_refs: A list of tuples, each containing ("worksheet", "B4"). If no worksheet (cell is on current sheet), returns ("no_sheet_referenced", "B4")
    """

    range_pattern = r"[A-Z$]+\d+:[A-Z$]+\d+" #parsing ranges got too hard simpler to just kill them. It makes the formula a bit messy — could dial this in more if needed. But fine for cell extraction
    formula_without_ranges = re.sub(range_pattern, '', formula)

    cell_pattern = r"(?:(?:'([^']+)'!))?\s*(?<!:)(\$?[A-Z]+\$?\d+)(?!:)"

    matches = re.findall(cell_pattern, formula_without_ranges)

    cell_refs = []
    for sheet, cell in matches:
        sheet = sheet if sheet else "no_sheet_referenced"
        cell = re.sub(r'\$', '', cell)
        cell_refs.append((sheet, cell))
    return cell_refs

formula = """=IFERROR(INDEX(XLOOKUP($C65,'Master Coverage Ratios'!$I$24:$AD$24,'Master Coverage Ratios'!$I$27:$AD$61),MATCH(H$59,'Master Coverage Ratios'!$B$27:$B$61,0,$C65:$C$66)),"Unavailable")"""
extract_cells(formula)


def cell_to_context(cell, cell_ws, formula_ws, cell_lookup, values_wb):
    """
    Find the context (row header title) for a specific cell

    Args:
        cell (str): e.g. "B3", no dollar signs

        cell_ws (str): title of the worksheet this cell exists in, if no sheet (current sheet) then will show "no_sheet_referenced"
        
        formula_ws (str): the title of the worksheet that the formula we are looking at is in. If the cell is no_sheet_referenced then we know that we should use the formula_ws to query it in cell_lookup

        cell_lookup(dict): A dict of worksheets, where each worksheet key corresponds to a dict of cell lookups. e.g. {"wksht1":{"B4":{stuff},"B3":{stuff}},"wksht2"...} see test.json for example.
    
    Returns:
        context (str): a string that describes the context of the cell (for use in prompt)
    """
    prefix = f"'{cell_ws}'!" #this is for showing LLM the exact cell reference


    if cell_ws == "no_sheet_referenced":
        cell_ws = formula_ws
        prefix = ''

    try:
        cell_data = cell_lookup[cell_ws][cell]
        context = f"*{prefix}{cell}* points to row: '{cell_data['row_descrip']}' in col: '{cell_data['col_descrip']}' in table: '{cell_data['title']}'"

    except KeyError:

        try:
            cell_data = values_wb[cell_ws][cell].value            

            context = f"*{prefix}{cell}* holds the value: {'[THIS CELL IS EMPTY]' if cell_data is None else cell_data}"

        except KeyError:
            print(f"ERROR. could not find *{cell}* in cell_lookup or in values_wb. Not good. Returning nothing.")
                
            context = ""

    return context

# RANGES
def extract_ranges(formula):
    """
    Take the formula and extract all the ranges

    Args:
        formula (str): the formula in the cell

    Returns
        range_refs (list): a list of all the range refs. Each range ref is a tuple which has worksheet ('worksheet', range)
    
    """
    range_pattern = r"(?:(?:'([^']+)'!))?\s*(\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+)"
    matches = re.findall(range_pattern, formula)
    
    range_refs = []
    for sheet, range_ref in matches:
        sheet = sheet if sheet else "no_sheet_referenced"
        range_ref = re.sub(r'\$', '', range_ref)
        range_refs.append((sheet, range_ref))
    return range_refs

def range_to_context(range_str, range_ws, formula_ws, cell_lookup):
    """
    From range, give the context of what this range means

    Args:
        range_str (str): "C34:D45" no $ please
        
        cell_ws (str): title of the worksheet that the range is in (for cell_lookup)
        
        formula_ws (str): title of the worksheet that the formula we are writing context for is in (for replacing)

        cell_lookup: dict of dicts where first dict is of worksheets, then you have your dict of cells. Access via cell_lookupo['worksheet']['B4'] and then you can find row_descrip col_descrip or title if you want. See test.json for example.

    Returns:
        context (str): a string that explains for an LLM what the context of the ranges is

    Ok let's think through this one. We have our range_str, range_ws, formlua_ws and cell lookupo. Maybe we can andlet his similarly

    first things first we get the range_ws set so that it deals with the no_sheet cases.

    Now we just have range_str, range_ws, cell_lookup. What do you want to happen if there is an error being thrown?

    Basically. Step 1 we try the happy path: which is to find the sheet in cell_lookup, then grab that sheet, then do our construction. There's no backup right? beacuse I don't want to return a value or anything like that. 

    Backup is just error rhandling

    """

    prefix = f"'{range_ws}'!"
    if range_ws == "no_sheet_referenced":
        range_ws = formula_ws
        prefix = ''
    
    if range_ws not in cell_lookup:
        print(f"range_ws '{range_ws}' not in cell_lookup.")
        return ""
    
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        row_descrips = {}
        for row in range(min_row, max_row + 1):
            key = f"row_{row}"
            reference_cell = f'{get_column_letter(min_col)}{row}'
            value = cell_lookup[range_ws][reference_cell]['row_descrip']
            row_descrips[key] = value

        col_descrips = {}
        for col in range(min_col, max_col + 1):
            key = f"col_{get_column_letter(col)}"
            reference_cell = f'{get_column_letter(col)}{min_row}'
            value = cell_lookup[range_ws][reference_cell]['col_descrip']
            col_descrips[key] = value

        first_cell = f"{get_column_letter(min_col)}{min_row}"
        
        range_title = cell_lookup[range_ws][first_cell]['title']
        
        context = f"<{prefix}{range_str}>*{range_str}* is defined by the following row and column descriptions: {str(row_descrips)} {str(col_descrips)} in the broader table '{range_title}'</{prefix}{range_str}>"

        return context
    
    except KeyError as e:
        print(f"RANGE_CONTEXT: KeyError encountered. Missing key: {e}")
        # print("Traceback:")
        # traceback.print_exc()
        return ""

    except Exception as e:
        print(f"RANGE_CONTEXT: An unexpected error occurred: {e}")
    
        return ""


def formula_context(formula, formula_ws, cell_lookup,values_wb):
    """
    Takes a formula and outputs a description of the context of the cells and ranges in the formula for an LLM to read
    
    Args:
        formula (str): the formula that you want to interpret
        
        formula_ws (str): title of the sheet the formula is in. 
        
        cell_lookup (dict): your workbook-level dictionary that has keys for each of the sheets, then the values of the sheets are a dict of cells with the context for each cell

        values_wb (dict)
        
    Returns:
        full_context (str): context of range explanations and cell explanations"""
    cell_explanations = []
    cell_refs = extract_cells(formula)  # this is going to return tuples ("ws", "B4")
    for cell_ws, cell in cell_refs:
        cell_context = cell_to_context(cell, cell_ws, formula_ws, cell_lookup, values_wb=values_wb)
        if cell_context:
            cell_explanations.append(cell_context)

    range_explanations = []
    ranges = extract_ranges(formula)  # also returns tuples
    for range_ws, r in ranges:
        range_context = range_to_context(range_str=r, range_ws=range_ws, formula_ws=formula_ws, cell_lookup=cell_lookup)
        if range_context:
            range_explanations.append(range_context)
    
    full_context = "\n".join(range_explanations + cell_explanations)
    return full_context.strip()

# # Test cases
# cell_lookup = {
#     "Example Worksheet": {
#     "A1": {"col_descrip": "Region", "row_descrip": "Header", "title": "Sales Data"},
#     "A2": {"col_descrip": "Region", "row_descrip": "North America", "title": "Sales Data"},
#     "A3": {"col_descrip": "Region", "row_descrip": "Europe", "title": "Sales Data"},
#     "A4": {"col_descrip": "Region", "row_descrip": "Asia", "title": "Sales Data"},
#     "B1": {"col_descrip": "Q1 Sales", "row_descrip": "Header", "title": "Sales Data"},
#     "B2": {"col_descrip": "Q1 Sales", "row_descrip": "North America", "title": "Sales Data"},
#     "B3": {"col_descrip": "Q1 Sales", "row_descrip": "Europe", "title": "Sales Data"},
#     "B4": {"col_descrip": "Q1 Sales", "row_descrip": "Asia", "title": "Sales Data"},
#     "C1": {"col_descrip": "Q2 Sales", "row_descrip": "Header", "title": "Sales Data"},
#     "C2": {"col_descrip": "Q2 Sales", "row_descrip": "North America", "title": "Sales Data"},
#     "C3": {"col_descrip": "Q2 Sales", "row_descrip": "Europe", "title": "Sales Data"},
#     "C4": {"col_descrip": "Q2 Sales", "row_descrip": "Asia", "title": "Sales Data"},
#     "D1": {"col_descrip": "Q3 Sales", "row_descrip": "Header", "title": "Sales Data"},
#     "D2": {"col_descrip": "Q3 Sales", "row_descrip": "North America", "title": "Sales Data"},
#     "D3": {"col_descrip": "Q3 Sales", "row_descrip": "Europe", "title": "Sales Data"},
#     "D4": {"col_descrip": "Q3 Sales", "row_descrip": "Asia", "title": "Sales Data"}
#     #     },
#     # "SECOND WORKSHEET": {
#     # "A1": {"col_descrip": "Region", "row_descrip": "Header", "title": "Sales Data"},
#     # "A2": {"col_descrip": "Region", "row_descrip": "North America", "title": "Sales Data"},
#     # "A3": {"col_descrip": "Region", "row_descrip": "Europe", "title": "Sales Data"},
#     # "A4": {"col_descrip": "Region", "row_descrip": "Asia", "title": "Sales Data"},
#     # "B1": {"col_descrip": "Q1 Sales", "row_descrip": "Header", "title": "Sales Data"},
#     # "B2": {"col_descrip": "Q1 Sales", "row_descrip": "North America", "title": "Sales Data"},
#     # "B3": {"col_descrip": "Q1 Sales", "row_descrip": "Europe", "title": "Sales Data"},
#     # "B4": {"col_descrip": "Q1 Sales", "row_descrip": "Asia", "title": "Sales Data"},
#     # "C1": {"col_descrip": "Q2 Sales", "row_descrip": "Header", "title": "Sales Data"},
#     # "C2": {"col_descrip": "Q2 Sales", "row_descrip": "North America", "title": "Sales Data"},
#     # "C3": {"col_descrip": "Q2 Sales", "row_descrip": "Europe", "title": "Sales Data"},
#     # "C4": {"col_descrip": "Q2 Sales", "row_descrip": "Asia", "title": "Sales Data"},
#     # "D1": {"col_descrip": "Q3 Sales", "row_descrip": "Header", "title": "Sales Data"},
#     # "D2": {"col_descrip": "Q3 Sales", "row_descrip": "North America", "title": "Sales Data"},
#     # "D3": {"col_descrip": "Q3 Sales", "row_descrip": "Europe", "title": "Sales Data"},
#     # "D4": {"col_descrip": "Q3 Sales", "row_descrip": "Asia", "title": "Sales Data"}
#     }
# }

# formula = """A1 + B2 + 'SECOND WORKSHEET'!$C3 + SUM($A1:$C2) + INDEX('SECOND WORKSHEET'!D3:D4, MATCH(A1, 'SECOND WORKSHEET'!A3:A4, 0)) + SUM($A1:A3) + D3 + 'SECOND WORKSHEET'!A1"""

# print(formula_context(formula, 'Example Worksheet', cell_lookup))