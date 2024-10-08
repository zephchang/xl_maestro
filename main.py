from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
import re
from openai import OpenAI
from dotenv import load_dotenv
import os
import json
import semantic_map
import parse

# Load the workbook_map.json file
with open('workbook_map.json', 'r') as f:
    workbook_map = json.load(f)

cell_lookup = semantic_map.semantic_map_workbook(workbook_map) #note this implicitly has the name of the spreadsheet in the workbook_map which tells semantic map to load that workbook from root.

# Load the workbook
formulas_wb = load_workbook('kc_big.xlsm', data_only=False)
values_wb = load_workbook('kc_big.xlsm', data_only=True)


# Save the result to test.json
with open('test.json', 'w') as f:
    json.dump(cell_lookup, f, indent=4)

print("Results saved to test.json")

print("CONTEXT FOR LLM\n", parse.formula_context(formula ="""=IFERROR(INDEX(XLOOKUP($C65,'Master Coverage Ratios'!$I$24:$AD$24,'Master Coverage Ratios'!$I$27:$AD$61),MATCH(H$59,'Master Coverage Ratios'!$B$27:$B$61,0,$C65:$C$66)),"Unavailable")""",formula_ws = "S9-13, 29-36 | Ratio Summaries",cell_lookup = cell_lookup, values_wb = values_wb))




# def parse_cell_reference(cell_ref):
#     # Use regex to split the cell reference into letters and numbers
#     match = re.match(r'([A-Za-z]+)(\d+)', cell_ref)
#     if match:
#         letters, numbers = match.groups()
#         return letters, int(numbers)
#     else:
#         raise ValueError(f"Invalid cell reference: {cell_ref}")

# def parse_cell_numbers(cell_ref):
#     # Use regex to split the cell reference into letters and numbers
#     match = re.match(r'([A-Za-z]+)(\d+)', cell_ref)
#     if match:
#         letters, numbers = match.groups()
#         # Convert column letters to number
#         column_number = sum((ord(char) - 64) * (26 ** i) for i, char in enumerate(reversed(letters.upper())))
#         return column_number, int(numbers)
#     else:
#         raise ValueError(f"Invalid cell reference: {cell_ref}")
    
# def parse_range_for_llm(range_start, range_end,formula_ws):
#     start_col, start_row = parse_cell_reference(range_start)
#     end_col, end_row = parse_cell_reference(range_end)

#     for row in value_ws.iter_rows(min_row=start_row, max_row = end_row, min_col=start_col, max_col=end_col):
#         for cell in row:
#             print(f"Cell {cell.coordinate}:")
#             print(f"   Value: {cell.value}")
#             print(f"   Data type: {cell.data_type}")
#             formula_cell = formula_ws[cell.coordinate] #testing committing from git directly more testing
#             if formula_cell.data_type == 'f':
#                 print(f"    HAS FORMULA: Formula: {formula_cell.value}")
#         print()

# def l_to_n(column_letter):
#     """
#     Convert a column letter (e.g., 'A', 'B', 'AA') to its corresponding column number (e.g., 1, 2, 27).
#     """
#     column_number = 0
#     for char in column_letter:
#         column_number = column_number * 26 + (ord(char.upper()) - ord('A') + 1)
#     return column_number

# def n_to_l(column_number):
#     """
#     Convert a column number (e.g., 1, 2, 27) to its corresponding column letter (e.g., 'A', 'B', 'AA').
#     """
#     column_letter = ""
#     while column_number > 0:
#         column_number, remainder = divmod(column_number - 1, 26)
#         column_letter = chr(65 + remainder) + column_letter
#     return column_letter

# def guess_cell_formula(cell, range_start, range_end, formula_ws):
#     cell_col, cell_row = parse_cell_reference(cell)
#     start_col_letter, start_row = parse_cell_reference(range_start)
#     end_col_letter, end_row = parse_cell_reference(range_end)

#     cell_col_number = l_to_n(cell_col)
#     start_col_number = l_to_n(start_col_letter)
#     end_col_number = l_to_n(end_col_letter)

#     workbook_purpose = "This workbook calculates and analyzes the financial performance of a business in several scenarios, focusing on metrics like revenue, gross margin, variable profit, and fixed costs. It also computes pre-tax profit, net margins, and cash flows for three individuals, taking into account tax rates, debt payments, and profit-sharing percentages."

#     col_headers = {} #for range end_col - start_col
#     for col_number in range(start_col_number,end_col_number+1):
#         cell_value = formula_ws.cell(row=start_row,column=col_number).value
#         col_headers[f'col_{n_to_l(col_number)}'] = cell_value

#     row_headers = {}
#     for row_number in range(start_row, end_row + 1):
#         cell_value = formula_ws.cell(row=row_number, column=start_col_number).value
#         row_headers[f'row_{row_number}'] = cell_value

#     cell_row_header = formula_ws.cell(row=cell_row,column=start_col_number).value
#     cell_col_header = formula_ws.cell(row=start_row,column=cell_col_number).value

#     prompt = f"""
#     You are a keen-eyed excel expert who is skilled at catching subtle formula errors or typos.
#     \n
#     You are looking at a workbook which has the following purpose: {workbook_purpose}
#     \n
#     The workbook has the following rows and collumns:

#     ROWS = {row_headers}

#     COLUMNS = {col_headers}

#     You are currently determining the correct formula for {cell}. {cell} is defined as {cell_row_header} (Row {cell_row}) in {cell_col_header} (Column {cell_col})

#     What would you expect {cell}'s formula to be? Answer in 2 sentences.
#     """
    
#     messages = [{"role": "user", "content": prompt}]
#     completion = client.chat.completions.create(
#         model="gpt-4o",
#         messages=messages
#     )
#     ai_guess = completion.choices[0].message.content
#     messages.append({"role":"assistant", "content":ai_guess})
#     messages.append({"role":"user", "content":f"The employee has attempted to write the formula, they wrote: {formula_ws.cell(row=cell_row, column=cell_col_number).value}\n\n\nDoes the employee's formula match with yours? If it doesn't match, why might that be? Does it need to be fixed? Respond YES it's fine to leave as is. or NO it needs to be fixed. Format you response as [Y] or [N]"})

#     evaluation = client.chat.completions.create(
#         model="gpt-4o",
#         messages=messages
#     )

#     evaluation = evaluation.choices[0].message.content
#     print(evaluation)

#     match = re.search(r'\[(Y|N)\]', evaluation)
#     if match:
#         verdict = match.group(1)
#     else:
#         verdict = "Unknown"
    
#     reasoning = f"AI GUESS: {ai_guess}\n\n EVALUATION: {evaluation}\n\nVERDICT: {verdict}\n\nPROMPT: {prompt}"

#     return verdict, reasoning

# def check_range(start_cell, end_cell, formula_ws):
#     # Parse start and end cells
#     start_col, start_row = parse_cell_reference(start_cell)
#     end_col, end_row = parse_cell_reference(end_cell)

#     # Convert column letters to numbers
#     start_col_num = l_to_n(start_col)
#     end_col_num = l_to_n(end_col)

#     # Iterate over the range
#     for row in range(start_row, end_row + 1):
#         for col_num in range(start_col_num, end_col_num + 1):
#             # Convert column number back to letter
#             col_letter = n_to_l(col_num)
#             cell = f"{col_letter}{row}"
#             # Check if the cell contains a formula
#             if formula_ws[cell].data_type == 'f':
#                 # Guess the formula for the cell
#                 verdict, reasoning = guess_cell_formula(cell, start_cell, end_cell, formula_ws)
#                 # Add the response as a comment
#                 cell_obj = formula_ws[cell]
#                 cell_obj.comment = Comment(text=reasoning, author="Maestro")
#                 if verdict == 'Y':
#                     cell_obj.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
#                 elif verdict == 'I':
#                     cell_obj.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
#                 elif verdict == 'N':
#                     cell_obj.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
#             else:
#                 pass

#             print(f"{cell} processed")

#         # Save the workbook after each row to balance between safety and performance
#         formula_wb.save('output.xlsx')

#     # Final save after processing all cells
#     formula_wb.save('output.xlsx')
#     print("Range check completed")
