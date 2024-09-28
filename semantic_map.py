from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
import json


def semantic_map_table(config, modify_dict): #look I don't love that we're using mutation, but we need some way to deal with overwriting so I prefer this than some kind of dictionary combining utility function
    workbook = config['workbook']
    worksheet = config['worksheet']
    col_descriptors = config['col_descriptors']
    row_descriptors = config['row_descriptors']
    check_cell_range = config['check_cell_range']
    table_title = config['table_title']
    
    value_wb = load_workbook(workbook, data_only=True)
    ws = value_wb[worksheet]

    col_start_col, col_start_row, col_end_col, col_end_row = range_boundaries(col_descriptors)
    row_start_col, row_start_row, row_end_col, row_end_row = range_boundaries(row_descriptors)

    cells_start_col, cells_start_row, cells_end_col, cells_end_row = range_boundaries(check_cell_range)

    col_headers = {} #this should be {1: "decription", 2: "description"}
    row_headers = {} #this is also {1: "decription", 2: "description"}


    for col in range(col_start_col, col_end_col+1):
        col_headers[col] = ws.cell(row=col_start_row, column=col).value or "NULL"
    
    for row in range(row_start_row, row_end_row+1):
        row_headers[row] = ws.cell(row=row, column=row_start_col).value or "NULL"
    
    for row in range(cells_start_row,cells_end_row+1):
        for col in range(cells_start_col, cells_end_col+1):
            modify_dict[f"{get_column_letter(col)}{row}"] = {"col_descrip":col_headers[col], "row_descrip":row_headers[row],"title":table_title}
        #should throw an error if col/row already exists
        
def semantic_map_workbook(workbook_map):
    workbook_tree = {}
    for worksheet in workbook_map["worksheets"]:
        worksheet_tree = {}
        for table in worksheet["tables"]:
            table_dic = {
                "workbook": workbook_map["wb_title"],
                "worksheet": worksheet["ws_title"],
                "table_title": table["title"],
                "col_descriptors": table["col_descriptors"],
                "row_descriptors": table["row_descriptors"],
                "check_cell_range": table["check_cell_range"]
            }
            semantic_map_table(table_dic, worksheet_tree)
        workbook_tree[worksheet["ws_title"]] = worksheet_tree
    return workbook_tree

        #ok so in theory, at this point we've gone through every worksheet and for each worksheet gone through every table and for each table constructed a dict and for each dict we added all of the cells to the semantic_map_table